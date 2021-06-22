import json
from datetime import datetime
from io import BytesIO

import pandas as pd
from flask import Response
from flask_apispec import MethodResource
from flask_apispec import use_kwargs, doc
from flask_jwt_extended import jwt_required
from flask_restful import Resource
from openpyxl.styles import PatternFill, Color, Font
from webargs import fields

from decorator.catch_exception import catch_exception
from decorator.log_request import log_request
from decorator.verify_admin_access import verify_admin_access
from utils.serializer import Serializer


class ExtractCompanies(MethodResource, Resource):

    db = None

    def __init__(self, db):
        self.db = db

    @log_request
    @doc(tags=['company'],
         description='Extract an excel file of the companies',
         responses={
             "200": {},
             "422": {"description": "Object not found"},
         })
    @use_kwargs({
        'name': fields.Str(required=False),
        'format': fields.Str(required=False),
        'include_address': fields.Bool(required=False),
        'include_email': fields.Bool(required=False),
        'include_phone': fields.Bool(required=False),
        'include_workforce': fields.Bool(required=False),
        'include_taxonomy': fields.Bool(required=False),
        'ecosystem_role': fields.DelimitedList(fields.Str(), required=False),
        'entity_type': fields.DelimitedList(fields.Str(), required=False),
        'startup_only': fields.Bool(required=False),
        'corebusiness_only': fields.Bool(required=False),
        'taxonomy_values': fields.DelimitedList(fields.Str(), required=False),
    }, location="query")
    @jwt_required
    @verify_admin_access
    @catch_exception
    def get(self, **kwargs):  # pylint: disable=too-many-branches

        # Manage global data

        companies = Serializer.serialize(self.db.get_filtered_companies(kwargs), self.db.tables["Company"])
        company_ids = [c["id"] for c in companies] \
            if len(companies) < self.db.get_count(self.db.tables["Company"]) else None

        df = pd.DataFrame(companies)
        df = df.add_prefix('Global|')

        # Manage addresses

        if 'include_address' in kwargs and kwargs['include_address'] is True:
            if company_ids is not None:
                addresses = self.db.get(self.db.tables["Company_Address"], {"company_id": company_ids})
            else:
                addresses = self.db.get(self.db.tables["Company_Address"])
            addresses = Serializer.serialize(addresses, self.db.tables["Company_Address"])
            addresses = pd.DataFrame(addresses)
            addresses = addresses.add_prefix('Address|')

            if len(addresses) > 0:
                df = df.merge(addresses, left_on='Global|id', right_on='Address|company_id', how='left')
                df = df.drop(['Address|id', 'Address|company_id'], axis=1)

        # Manage email addresses from contacts

        if 'include_email' in kwargs and kwargs['include_email'] is True:
            if company_ids is not None:
                contacts = self.db.get(self.db.tables["CompanyContact"], {
                    "company_id": company_ids,
                    "type": "EMAIL ADDRESS",
                })
            else:
                contacts = self.db.get(self.db.tables["CompanyContact"], {
                    "type": "EMAIL ADDRESS",
                })
            contacts = Serializer.serialize(contacts, self.db.tables["CompanyContact"])
            contacts = pd.DataFrame(contacts)
            contacts = contacts.add_prefix('Email|')

            if len(contacts) > 0:
                df = df.merge(contacts, left_on='Global|id', right_on='Email|company_id', how='left')
                df = df.drop(['Email|id', 'Email|company_id', 'Email|type'], axis=1)

        # Manage phone numbers from contacts

        if 'include_phone' in kwargs and kwargs['include_phone'] is True:
            if company_ids is not None:
                contacts = self.db.get(self.db.tables["CompanyContact"], {
                    "company_id": company_ids,
                    "type": "PHONE NUMBER",
                })
            else:
                contacts = self.db.get(self.db.tables["CompanyContact"], {
                    "type": "PHONE NUMBER",
                })
            contacts = Serializer.serialize(contacts, self.db.tables["CompanyContact"])
            contacts = pd.DataFrame(contacts)
            contacts = contacts.add_prefix('Phone|')

            if len(contacts) > 0:
                df = df.merge(contacts, left_on='Global|id', right_on='Phone|company_id', how='left')
                df = df.drop(['Phone|id', 'Phone|company_id', 'Phone|type'], axis=1)

        # Manage workforces

        if 'include_workforce' in kwargs and kwargs['include_workforce'] is True:
            workforces = self.db.get_latest_workforce(company_ids)
            workforces = Serializer.serialize(workforces, self.db.tables["Workforce"])
            workforces = pd.DataFrame(workforces)
            workforces = workforces.add_prefix('Workforce|')

            if len(workforces) > 0:
                df = df.merge(workforces, left_on='Global|id', right_on='Workforce|company', how='left')
                df = df.drop(['Workforce|id', 'Workforce|company'], axis=1)

        # Manage taxonomy

        if 'include_taxonomy' in kwargs and kwargs['include_taxonomy'] is True:
            df = self.include_taxonomy(company_ids, df)

        # Prepare final export

        if 'format' not in kwargs or kwargs['format'] != "json":
            filename = f"Export - Companies - {datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.xlsx"

            res = Response(
                self.prepare_xlsx(df),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment;filename={filename}"}
            )

            return res
        return json.loads(df.to_json(orient="records")), "200 "

    def include_taxonomy(self, company_ids, df):
        assignments = self.db.get(self.db.tables["TaxonomyAssignment"],
                                  {"company": company_ids} if company_ids is not None else {})

        categories = self.db.get(self.db.tables["TaxonomyCategory"])
        values = self.db.get(self.db.tables["TaxonomyValue"])
        category_hierarchy = self.db.get(self.db.tables["TaxonomyCategoryHierarchy"])
        value_hierarchy = self.db.get(self.db.tables["TaxonomyValueHierarchy"])

        parent_categories = [c.parent_category for c in category_hierarchy]
        child_categories = [c.name for c in categories if c.name not in parent_categories]
        child_values = [v.id for v in values if v.category in child_categories]
        assignments = [a for a in assignments if a.taxonomy_value in child_values]

        taxonomy = {}

        for a in assignments:
            value = [v for v in values if v.id == a.taxonomy_value][0]
            self.add_taxonomy_value(taxonomy, a.company, value)

            while value.id in [vh.child_value for vh in value_hierarchy]:
                vh = [vh for vh in value_hierarchy if vh.child_value == value.id][0]
                value = [v for v in values if v.id == vh.parent_value][0]
                self.add_taxonomy_value(taxonomy, a.company, value)

        taxonomy = pd.DataFrame(taxonomy).transpose()
        taxonomy = taxonomy.applymap(lambda x: ";".join(x) if isinstance(x, list) else x)
        taxonomy = taxonomy.reset_index()
        taxonomy = taxonomy.add_prefix('Taxonomy|')

        if len(taxonomy) > 0:
            taxonomy = pd.DataFrame(taxonomy)
            df = df.merge(taxonomy, left_on='Global|id', right_on='Taxonomy|index', how='left')
            df = df.drop(['Taxonomy|index'], axis=1)

        return df

    def prepare_xlsx(self, df):
        # Clean DF

        df = df.drop('Global|id', axis=1)
        df = df.rename({'Global|rscl_number': "rscl_number"}, axis=1)
        df = df.set_index("rscl_number")
        df.columns = df.columns.str.split('|', expand=True)

        # Build the XLS file

        xlsx = BytesIO()
        writer = pd.ExcelWriter(xlsx, engine='openpyxl')  # pylint: disable=abstract-class-instantiated
        df.to_excel(writer, startrow=0, sheet_name="Companies")
        sheet = writer.sheets["Companies"]
        self.apply_style(sheet)
        writer.save()
        xlsx.seek(0)

        return xlsx

    @staticmethod
    def apply_style(sheet):
        for cell in sheet["A:A"]:
            cell.fill = PatternFill(fgColor=Color("8fddff"), fill_type="solid")
            cell.font = Font(bold=True)
        for cell in sheet["1:1"]:
            cell.fill = PatternFill(fgColor=Color("ffa8b0"), fill_type="solid")
            cell.font = Font(bold=True)
        for cell in sheet["2:2"]:
            cell.fill = PatternFill(fgColor=Color("ffa8b0"), fill_type="solid")
            cell.font = Font(bold=True)

        dims = {}

        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = min(30, value)

        sheet.column_dimensions["A"].width = 15

    @staticmethod
    def add_taxonomy_value(table, company, value):
        if company not in table:
            table[company] = {}

        if value.category not in table[company]:
            table[company][value.category] = []

        if value.name not in table[company][value.category]:
            table[company][value.category].append(value.name)
