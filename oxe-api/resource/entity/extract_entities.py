import json
from datetime import datetime
from io import BytesIO

import pandas as pd
from flask import Response
from flask_apispec import MethodResource
from flask_apispec import use_kwargs, doc
from flask_jwt_extended import jwt_required
from flask_restful import Resource
from openpyxl.styles import PatternFill, Color, Font
from webargs import fields

from decorator.catch_exception import catch_exception
from decorator.log_request import log_request
from decorator.verify_admin_access import verify_admin_access
from utils.serializer import Serializer
from sqlalchemy import or_

class ExtractEntities(MethodResource, Resource):

    db = None

    def __init__(self, db):
        self.db = db

    @log_request
    @doc(tags=['entity'],
         description='Extract the entities into a file',
         responses={
             "200": {},
         })
    @use_kwargs({
        'name': fields.Str(required=False),
        'format': fields.Str(required=False),
        'include_user': fields.Bool(required=False),
        'include_address': fields.Bool(required=False),
        'include_email': fields.Bool(required=False),
        'include_phone': fields.Bool(required=False),
        'include_workforce': fields.Bool(required=False),
        'include_taxonomy': fields.Bool(required=False),
        'include_authorisation_by_approved_signatory': fields.Bool(required=False),
        'ecosystem_role': fields.DelimitedList(fields.Str(), required=False),
        'entity_type': fields.DelimitedList(fields.Str(), required=False),
        'startup_only': fields.Bool(required=False),
        'corebusiness_only': fields.Bool(required=False),
        'taxonomy_values': fields.DelimitedList(fields.Int(), required=False),
    }, location="query")
    @jwt_required
    @verify_admin_access
    @catch_exception  # pylint: disable=too-many-branches,too-many-statements
    def get(self, **kwargs):  # noqa: MC0001

        # Manage global data

        entities = Serializer.serialize(self.db.get_filtered_entities(kwargs).all(), self.db.tables["Entity"])
        entity_ids = [c["id"] for c in entities] \
            if len(entities) < self.db.get_count(self.db.tables["Entity"]) else None

        df = pd.DataFrame(entities)
        df = df.add_prefix('Global|')

        # Manage addresses

        if 'include_address' in kwargs and kwargs['include_address'] is True:
            if entity_ids is not None:
                addresses = self.db.get(self.db.tables["EntityAddress"], {"entity_id": entity_ids})
            else:
                addresses = self.db.get(self.db.tables["EntityAddress"])
            addresses = Serializer.serialize(addresses, self.db.tables["EntityAddress"])
            addresses = pd.DataFrame(addresses)
            addresses = addresses.add_prefix('Address|')

            if len(addresses) > 0:
                df = df.merge(addresses, left_on='Global|id', right_on='Address|entity_id', how='left')
                df = df.drop(['Address|id', 'Address|entity_id'], axis=1)

        # Manage users

        if 'include_user' in kwargs and kwargs['include_user'] is True:
            if entity_ids is not None:
                user_assignments = self.db.get(self.db.tables["UserEntityAssignment"], {
                    "entity_id": entity_ids,
                })
            else:
                user_assignments = self.db.get(self.db.tables["UserEntityAssignment"])

            users = self.db.get(
                self.db.tables["User"],
                {"id": list(set([a.user_id for a in user_assignments]))},  # pylint: disable=consider-using-set-comprehension
                ["id", "email"]
            )

            user_assignments = Serializer.serialize(user_assignments, self.db.tables["UserEntityAssignment"])

            if len(users) > 0:
                user_assignments = pd.DataFrame(user_assignments)
                users = pd.DataFrame(users)
                users = user_assignments.merge(users, left_on='user_id', right_on='id', how='left')
                users = users.drop(labels=["id"], axis=1)
                users = users.add_prefix('User|')

                df = df.merge(users, left_on='Global|id', right_on='User|entity_id', how='left')
                df = df.drop(['User|user_id', 'User|entity_id'], axis=1)

        # Manage email addresses from contacts

        if 'include_email' in kwargs and kwargs['include_email'] is True:
            if entity_ids is not None:
                contacts = self.db.get(self.db.tables["EntityContact"], {
                    "entity_id": entity_ids,
                    "type": "EMAIL ADDRESS",
                })
            else:
                contacts = self.db.get(self.db.tables["EntityContact"], {
                    "type": "EMAIL ADDRESS",
                })
            contacts = Serializer.serialize(contacts, self.db.tables["EntityContact"])
            contacts = pd.DataFrame(contacts)
            contacts = contacts.add_prefix('Email|')

            if len(contacts) > 0:
                df = df.merge(contacts, left_on='Global|id', right_on='Email|entity_id', how='left')
                df = df.drop(['Email|id', 'Email|entity_id', 'Email|type'], axis=1)

        # Manage phone numbers from contacts

        if 'include_phone' in kwargs and kwargs['include_phone'] is True:
            if entity_ids is not None:
                contacts = self.db.get(self.db.tables["EntityContact"], {
                    "entity_id": entity_ids,
                    "type": "PHONE NUMBER",
                })
            else:
                contacts = self.db.get(self.db.tables["EntityContact"], {
                    "type": "PHONE NUMBER",
                })
            contacts = Serializer.serialize(contacts, self.db.tables["EntityContact"])
            contacts = pd.DataFrame(contacts)
            contacts = contacts.add_prefix('Phone|')

            if len(contacts) > 0:
                df = df.merge(contacts, left_on='Global|id', right_on='Phone|entity_id', how='left')
                df = df.drop(['Phone|id', 'Phone|entity_id', 'Phone|type'], axis=1)

        # Manage workforces

        if 'include_workforce' in kwargs and kwargs['include_workforce'] is True:
            workforces = self.db.get_latest_workforce(entity_ids)
            workforces = Serializer.serialize(workforces, self.db.tables["Workforce"])
            workforces = pd.DataFrame(workforces)
            workforces = workforces.add_prefix('Workforce|')

            if len(workforces) > 0:
                df = df.merge(workforces, left_on='Global|id', right_on='Workforce|entity_id', how='left')
                df = df.drop(['Workforce|id', 'Workforce|entity_id'], axis=1)

        # Manage taxonomy

        if 'include_taxonomy' in kwargs and kwargs['include_taxonomy'] is True:
            df = self.include_taxonomy(entity_ids, df)

        # Manage Authorisation by Approved Signatory
        if 'include_authorisation_by_approved_signatory' in kwargs and kwargs['include_authorisation_by_approved_signatory'] is True:
            vat_numbers = [c["vat_number"] for c in entities] \
                if len(entities) <= self.db.get_count(self.db.tables["Entity"]) else None

            if vat_numbers is not None:
                documents = self.db.session.query(self.db.tables["Document"])

                vat_filters = []
                for vat_number in vat_numbers:
                    vat_filters.append(self.db.tables["Document"].filename.like(f"%{vat_number}%"))

                # * Spread Operator in Python
                documents = documents.filter(or_(*vat_filters)).all()

                documents = Serializer.serialize(documents, self.db.tables["Document"])

                # documents = pd.DataFrame(documents)
                # documents = documents.add_prefix('Signatory Approval|filename|')

                if len(documents) > 0:
                    # @TODO: Improve this:
                    # Find away to merge using SQL "like" operator:
                    #  - self.db.tables["Document"].filename.like(f"%{vat_number}%")
                    #
                    # df = df.merge(documents, left_on='Global|vat_number', right_on='Document|filename', how='left')
                    # df = df.drop(['Document|X', 'Document|Y', '...'], axis=1)

                    # @TODO Bad complexity, need improvement.
                    df['Signatory Approval|filename'] = pd.Series(dtype='object')
                    for i, vat in df['Global|vat_number'].items():
                        if vat is None or len(vat) < 1:
                            continue
                        for doc in documents:
                            filename = doc['filename']
                            # if vat_number in filename: "entity_registration_approval_{user_id}_{vat_number}.pdf"
                            if vat in filename:
                                df.loc[i, 'Signatory Approval|filename'] = filename
                                break

        # Prepare final export

        if 'format' not in kwargs or kwargs['format'] != "json":
            filename = f"Export - entities - {datetime.now().strftime('%Y-%m-%d')}.xlsx"

            res = Response(
                self.prepare_xlsx(df),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment;filename={filename}"}
            )

            return res
        return json.loads(df.to_json(orient="records")), "200 "

    def include_taxonomy(self, entity_ids, df):
        assignments = self.db.get(self.db.tables["TaxonomyAssignment"],
                                  {"entity": entity_ids} if entity_ids is not None else {})

        categories = self.db.get(self.db.tables["TaxonomyCategory"])
        values = self.db.get(self.db.tables["TaxonomyValue"])
        category_hierarchy = self.db.get(self.db.tables["TaxonomyCategoryHierarchy"])
        value_hierarchy = self.db.get(self.db.tables["TaxonomyValueHierarchy"])

        parent_categories = [c.parent_category for c in category_hierarchy]
        child_categories = [c.name for c in categories if c.name not in parent_categories]
        child_values = [v.id for v in values if v.category in child_categories]
        assignments = [a for a in assignments if a.taxonomy_value_id in child_values]

        taxonomy = {}

        for a in assignments:
            value = [v for v in values if v.id == a.taxonomy_value_id][0]
            self.add_taxonomy_value(taxonomy, a.entity_id, value)

            while value.id in [vh.child_value for vh in value_hierarchy]:
                vh = [vh for vh in value_hierarchy if vh.child_value == value.id][0]
                value = [v for v in values if v.id == vh.parent_value][0]
                self.add_taxonomy_value(taxonomy, a.entity_id, value)

        taxonomy = pd.DataFrame(taxonomy).transpose()
        taxonomy = taxonomy.applymap(lambda x: ";".join(x) if isinstance(x, list) else x)
        taxonomy = taxonomy.reset_index()
        taxonomy = taxonomy.add_prefix('Taxonomy|')

        if len(taxonomy) > 0:
            taxonomy = pd.DataFrame(taxonomy)
            df = df.merge(taxonomy, left_on='Global|id', right_on='Taxonomy|index', how='left')
            df = df.drop(['Taxonomy|index'], axis=1)

        return df

    def prepare_xlsx(self, df):
        # Clean DF


        df = df.rename({'Global|id': "Entity_ID"}, axis=1)
        df = df.set_index("Entity_ID")
        df = df.rename({'Global|image': "Global|image_id"}, axis=1)
        df = df.drop('Global|description', axis=1)
        df = df.drop('Global|creation_date', axis=1)
        df = df.drop('Global|is_startup', axis=1)
        df = df.drop('Global|is_cybersecurity_core_business', axis=1)
        df = df.drop('Global|linkedin_url', axis=1)
        df = df.drop('Global|twitter_url', axis=1)
        df = df.drop('Global|youtube_url', axis=1)
        df = df.drop('Global|discord_url', axis=1)
        df = df.drop('Global|sync_id', axis=1)
        df = df.drop('Global|sync_node', axis=1)
        df = df.drop('Global|sync_global', axis=1)
        df = df.drop('Global|sync_address', axis=1)
        df = df.drop('Global|sync_status', axis=1)
        df = df.drop('Global|legal_status', axis=1)
        df = df.drop('Global|headline', axis=1)

        df = df.drop('Address|number', axis=1)
        df = df.drop('Address|administrative_area', axis=1)
        df = df.drop('Address|latitude', axis=1)
        df = df.drop('Address|longitude', axis=1)

        df = df.drop('User|department', axis=1)
        df = df.drop('User|seniority_level', axis=1)
        df = df.drop('User|email', axis=1)

        df = df.drop('Email|representative', axis=1)
        df = df.drop('Email|name', axis=1)
        df = df.drop('Email|value', axis=1)
        df = df.drop('Email|department', axis=1)
        df = df.drop('Email|user_id', axis=1)


        df = df.drop({'Global|trade_register_number'}, axis=1)
        df.columns = df.columns.str.split('|', expand=True)

        # Build the XLS file

        xlsx = BytesIO()
        writer = pd.ExcelWriter(xlsx, engine='openpyxl')  # pylint: disable=abstract-class-instantiated
        df.to_excel(writer, startrow=0, sheet_name="Entities")
        sheet = writer.sheets["Entities"]
        self.apply_style(sheet)
        writer.save()
        xlsx.seek(0)

        return xlsx

    @staticmethod
    def apply_style(sheet):
        for cell in sheet["A:A"]:
            cell.fill = PatternFill(fgColor=Color("8fddff"), fill_type="solid")
            cell.font = Font(bold=True)
        for cell in sheet["1:1"]:
            cell.fill = PatternFill(fgColor=Color("ffa8b0"), fill_type="solid")
            cell.font = Font(bold=True)
        for cell in sheet["2:2"]:
            cell.fill = PatternFill(fgColor=Color("ffa8b0"), fill_type="solid")
            cell.font = Font(bold=True)

        dims = {}

        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = min(30, value)

        sheet.column_dimensions["A"].width = 15

    @staticmethod
    def add_taxonomy_value(table, entity, value):
        if entity not in table:
            table[entity] = {}

        if value.category not in table[entity]:
            table[entity][value.category] = []

        if value.name not in table[entity][value.category]:
            table[entity][value.category].append(value.name)
