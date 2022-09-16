from datetime import datetime
from io import BytesIO

import pandas as pd
from flask import Response
from flask_apispec import MethodResource
from flask_apispec import use_kwargs, doc
from flask_jwt_extended import jwt_required
from flask_restful import Resource
from openpyxl.styles import PatternFill, Color, Font
from webargs import fields
from sqlalchemy.orm.exc import NoResultFound

from decorator.catch_exception import catch_exception
from decorator.log_request import log_request
from decorator.verify_admin_access import verify_admin_access
from utils.serializer import Serializer
from utils.speadsheet import clean_html


class ExtractForm(MethodResource, Resource):

    db = None

    def __init__(self, db):
        self.db = db

    @log_request
    @doc(tags=['form'],
         description='Extract the form and its answers into a file',
         responses={
             "200": {},
             "422": {"description": "Object not found"},
         })
    @use_kwargs({
        'id': fields.Int(required=True),
        'format': fields.Str(required=False, validate=lambda x: x in ['xlsx', 'json', None]),
    }, location="query")
    @jwt_required
    @verify_admin_access
    @catch_exception
    def get(self, **kwargs):

        # Get form

        form = None

        try:
            form = self.db.session \
                .query(self.db.tables["Form"]) \
                .filter(self.db.tables["Form"].id == kwargs["id"]) \
                .one()
        except NoResultFound:
            return "", "422 Object not found"

        # Get form questions and answers

        questions = self.db.session \
            .query(self.db.tables["FormQuestion"]) \
            .filter(self.db.tables["FormQuestion"].form_id == form.id) \
            .order_by(self.db.tables["FormQuestion"].position) \
            .all()

        answers = self.db.session \
            .query(self.db.tables["FormAnswer"]) \
            .filter(self.db.tables["FormAnswer"].form_question_id.in_([q.id for q in questions])) \
            .all()

        user_ids = list(set([a.user_id for a in answers]))

        users = self.db.session \
            .query(self.db.tables["User"]) \
            .with_entities(self.db.tables["User"].id, self.db.tables["User"].email) \
            .filter(self.db.tables["User"].id.in_(user_ids)) \
            .all()

        df = pd.DataFrame(
            index=[q.id for q in questions],
            columns=[u[0] for u in users],
        )

        for a in answers:
            df[a.user_id][a.form_question_id] = a.value

        # Prepare final export

        if 'format' not in kwargs or kwargs['format'] != "json":
            filename = f"Export - entities - {datetime.now().strftime('%Y-%m-%d')}.xlsx"

            res = Response(
                self.prepare_xlsx(df, questions, users),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment;filename={filename}"}
            )

            return res

        data = dict()
        data["form"] = Serializer.serialize(form, self.db.tables["Form"])
        data["questions"] = Serializer.serialize(questions, self.db.tables["FormQuestion"])
        data["answers"] = df.to_dict(orient="dict")

        return data, "200 "

    def prepare_xlsx(self, df, questions, users):
        # Replace user IDs by user email

        df = df.rename(columns={u[0]: u[1] for u in users})

        # Replace question IDs by question value

        df.index = df.index.rename("Questions")
        df = df.reset_index(level=0)
        df["Questions"] = df["Questions"].map({q.id: q.value for q in questions})

        # Clean cells

        for c in df.columns:
            df[c] = df[c].apply(lambda x: clean_html(x))

        # Build the XLS file

        xlsx = BytesIO()
        writer = pd.ExcelWriter(xlsx, engine='openpyxl')  # pylint: disable=abstract-class-instantiated
        df.to_excel(writer, startrow=0, sheet_name="Form", index=False)
        sheet = writer.sheets["Form"]
        self.apply_style(sheet)
        writer.save()
        xlsx.seek(0)

        return xlsx

    @staticmethod
    def apply_style(sheet):
        for cell in sheet["A:A"]:
            cell.fill = PatternFill(fgColor=Color("8fddff"), fill_type="solid")
            cell.font = Font(bold=True)
        for cell in sheet["1:1"]:
            cell.fill = PatternFill(fgColor=Color("ffa8b0"), fill_type="solid")
            cell.font = Font(bold=True)

        dims = {}

        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                    cell.alignment = cell.alignment.copy(alignment=True)
        for col, _ in dims.items():
            sheet.column_dimensions[col].width = 30

        sheet.column_dimensions["A"].width = 50
